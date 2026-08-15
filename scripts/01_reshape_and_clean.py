"""
Phase 1 - Data Foundation Pipeline
==================================
Input : data/Sales_Data.xlsx   (তোমার original wide-format sales file)
Output: outputs/Sales_Data_Cleaned_Phase1.xlsx  (5 sheets)
        outputs/sales_long_format.csv           (raw tidy/long format)
        outputs/clean_activity_level.csv        (modeling-ready base, sparse rows dropped)
        outputs/customer_master.csv             (deduplicated customer-level table)

চালানোর নিয়ম (VS Code terminal থেকে):
    python scripts/01_reshape_and_clean.py
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 0. PATHS  — দরকার হলে এখানে ফাইলের নাম/লোকেশন বদলাও
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
INPUT_FILE = PROJECT_ROOT / "data" / "Sales_Data.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

# এই কলামগুলো meta-data (brand block না)
META_COLS = [
    "ID", "Sales Person (Sales Team)", "Coordinators",
    "Customer Name", "Customer Group", "Zone",
]
TAIL_COLS = [
    "Total Yearly Target 26", "Total Achievement jan to june ", "Achievment%",
]


# ---------------------------------------------------------------------------
# 1. WIDE -> LONG (TIDY) FORMAT
#    প্রতিটা brand-এর 3টা column (Target, ACH, ACH%) কে আলাদা row-এ নিয়ে আসা
# ---------------------------------------------------------------------------
def reshape_to_long(df: pd.DataFrame) -> pd.DataFrame:
    cols = list(df.columns)
    brand_block = [c for c in cols if c not in META_COLS and c not in TAIL_COLS]

    assert len(brand_block) % 3 == 0, "Brand columns 3-er multiple na -- sheet structure check koro"

    records = []
    for i in range(0, len(brand_block), 3):
        brand_col, ach_col, achpct_col = brand_block[i], brand_block[i + 1], brand_block[i + 2]
        tmp = df[META_COLS + [brand_col, ach_col, achpct_col]].copy()
        tmp.columns = META_COLS + ["Target", "Achievement", "Achievement_pct"]
        tmp["Brand"] = brand_col.strip()
        records.append(tmp)

    long_df = pd.concat(records, ignore_index=True)
    return long_df[META_COLS + ["Brand", "Target", "Achievement", "Achievement_pct"]]


# ---------------------------------------------------------------------------
# 2. DATA QUALITY FLAGS
# ---------------------------------------------------------------------------
def add_flags(long_df: pd.DataFrame) -> pd.DataFrame:
    df = long_df.copy()
    df["has_target"] = df["Target"].notna() & (df["Target"] > 0)
    df["has_achievement"] = df["Achievement"].notna() & (df["Achievement"] > 0)
    df["activity_flag"] = np.where(df["has_target"] | df["has_achievement"], "has_activity", "no_activity")
    df["target_gap_flag"] = (~df["has_target"]) & df["has_achievement"]      # sale hoyeche, target set hoyni
    df["negative_value_flag"] = (df["Target"] < 0) | (df["Achievement"] < 0)
    df["Achievement_pct_clean"] = np.where(
        df["has_target"], df["Achievement"].fillna(0) / df["Target"], np.nan
    )
    return df


# ---------------------------------------------------------------------------
# 3. CUSTOMER MASTER TABLE
# ---------------------------------------------------------------------------
def build_customer_master(long_df: pd.DataFrame) -> pd.DataFrame:
    master = long_df.drop_duplicates(subset=["ID"])[
        ["ID", "Customer Name", "Sales Person (Sales Team)", "Coordinators", "Customer Group", "Zone"]
    ].reset_index(drop=True)

    totals = long_df.groupby("ID").agg(
        Total_Target=("Target", "sum"),
        Total_Achievement=("Achievement", "sum"),
        N_brands_active=("activity_flag", lambda x: (x == "has_activity").sum()),
    ).reset_index()

    master = master.merge(totals, on="ID", how="left")
    master["Overall_Achievement_pct"] = np.where(
        master["Total_Target"] > 0, master["Total_Achievement"] / master["Total_Target"], np.nan
    )
    master["is_zero_achiever"] = (master["Total_Target"] > 0) & (master["Total_Achievement"].fillna(0) == 0)
    return master


# ---------------------------------------------------------------------------
# 4. DATA QUALITY REPORT
# ---------------------------------------------------------------------------
def build_quality_report(long_df: pd.DataFrame, master: pd.DataFrame) -> pd.DataFrame:
    rows = [
        ("Total customer-brand combinations", len(long_df)),
        ("Rows with real activity (target or achievement)", (long_df["activity_flag"] == "has_activity").sum()),
        ("Rows: sale exists but NO target set (target gap)", long_df["target_gap_flag"].sum()),
        ("Value of untargeted sales (Taka)", long_df.loc[long_df["target_gap_flag"], "Achievement"].sum()),
        ("Rows with negative values", long_df["negative_value_flag"].sum()),
        ("Total unique customers", master["ID"].nunique()),
        ("Customers with target but zero achievement", master["is_zero_achiever"].sum()),
        ("Target value stuck in zero-achievers (Taka)", master.loc[master["is_zero_achiever"], "Total_Target"].sum()),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


# ---------------------------------------------------------------------------
# 5. EXCEL EXPORT (formatted, multi-sheet)
# ---------------------------------------------------------------------------
def write_sheet(wb: Workbook, name: str, df: pd.DataFrame):
    ws = wb.create_sheet(name)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            if r_idx == 1:
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
    for i, col in enumerate(df.columns, 1):
        try:
            maxlen = max(df[col].astype(str).map(len).max(), len(str(col)))
        except Exception:
            maxlen = 15
        ws.column_dimensions[get_column_letter(i)].width = min(max(maxlen + 2, 10), 45)
    ws.freeze_panes = "A2"


def export_excel(report_df, target_gap_list, zero_achievers, master, clean_activity, out_path: Path):
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Data_Quality_Report", report_df)
    write_sheet(wb, "Target_Gap_List", target_gap_list)
    write_sheet(wb, "Zero_Achievers", zero_achievers)
    write_sheet(wb, "Customer_Master", master)
    write_sheet(wb, "Clean_Activity_Level", clean_activity)
    wb.save(out_path)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    print(f"Reading: {INPUT_FILE}")
    df = pd.read_excel(INPUT_FILE)

    long_df = reshape_to_long(df)
    long_df = add_flags(long_df)
    master = build_customer_master(long_df)
    report_df = build_quality_report(long_df, master)

    clean_activity = long_df[long_df["activity_flag"] == "has_activity"].copy()
    target_gap_list = long_df[long_df["target_gap_flag"]].sort_values("Achievement", ascending=False)[
        ["Customer Name", "Sales Person (Sales Team)", "Zone", "Brand", "Achievement"]
    ]
    zero_achievers = master[master["is_zero_achiever"]].sort_values("Total_Target", ascending=False)[
        ["Customer Name", "Sales Person (Sales Team)", "Zone", "Total_Target"]
    ]

    # CSV exports (handy for quick pandas re-use later, e.g. segmentation script)
    long_df.to_csv(OUTPUT_DIR / "sales_long_format.csv", index=False)
    clean_activity.to_csv(OUTPUT_DIR / "clean_activity_level.csv", index=False)
    master.to_csv(OUTPUT_DIR / "customer_master.csv", index=False)

    # Excel export
    export_excel(
        report_df, target_gap_list, zero_achievers, master, clean_activity,
        OUTPUT_DIR / "Sales_Data_Cleaned_Phase1.xlsx",
    )

    print("\n=== DATA QUALITY REPORT ===")
    print(report_df.to_string(index=False))
    print(f"\nDone. Files saved in: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
